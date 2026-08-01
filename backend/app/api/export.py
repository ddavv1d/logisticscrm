"""Экспорт списка контейнеров в Excel (openpyxl) — выгружает текущий фильтр.

Деньги (доход/маржа/остаток) добавляются ТОЛЬКО для owner/accountant
(RBAC как в UI — менеджеру нельзя выгружать маржу в обход гейта).
Маршрут (цепочка плеч) — всем.
"""

import io
from decimal import Decimal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import current_user
from app.core.db import get_session
from app.domain.locations import LOCATION_TITLE_RU
from app.domain.stages import stage_meta
from app.models.models import User
from app.services.container_service import build_container_query, days_on_stage, is_stuck

router = APIRouter(prefix="/api/export", tags=["export"])

_BASE_HEAD = ["Ref №", "Контейнер №", "Тип", "Клиент", "Маршрут", "Стадия", "Дней на стадии", "Застрял"]
_MONEY_HEAD = ["Доход USD", "Расход USD", "Маржа USD", "Оплачено USD", "Остаток USD", "Просрочка"]


def _loc(code: str) -> str:
    return LOCATION_TITLE_RU.get(code, code)


def _route(c) -> str:
    """Цепочка плеч: Поти→Тбилиси→Баку→… (реальные плечи контейнера)."""
    legs = sorted(c.legs, key=lambda x: x.seq)
    if not legs:
        return _loc(c.origin_location) + " → " + _loc(c.dest_location)
    pts = [_loc(legs[0].from_location)] + [_loc(le.to_location) for le in legs]
    return " → ".join(pts)


def _money(c):
    """Считаем деньги в USD из charge_line.amount_usd (мультивалюта уже сведена)."""
    inc = exp = paid_inc = Decimal(0)
    for ch in c.charges:
        au = Decimal(str(ch.amount_usd or 0))
        if ch.kind == "income":
            inc += au
            # оплаченная доля дохода в USD (paid_amount * rate)
            rate = Decimal(str(ch.rate_to_usd or 1))
            paid_inc += Decimal(str(ch.paid_amount or 0)) * rate
        else:
            exp += au
    outstanding = inc - paid_inc
    return {
        "income": float(inc), "expense": float(exp), "margin": float(inc - exp),
        "paid": float(paid_inc), "outstanding": float(outstanding),
    }


@router.get("/containers.xlsx")
async def export_containers(
    session: AsyncSession = Depends(get_session),
    user: User = Depends(current_user),
    search: str | None = None,
    stage: str | None = None,
    client_id: int | None = None,
    chip: str | None = None,
    status_: str = Query("active", alias="status"),
):
    q = build_container_query(
        search=search, stage=stage, client_id=client_id, chip=chip, status=status_
    )
    rows = await session.execute(q)
    containers = list(rows.scalars().unique().all())
    if chip == "stuck":
        containers = [c for c in containers if is_stuck(c)]

    show_money = user.can_see_money  # RBAC: деньги только money-ролям

    wb = Workbook()
    ws = wb.active
    ws.title = "Контейнеры"
    head = _BASE_HEAD + (_MONEY_HEAD if show_money else [])
    ws.append(head)
    head_fill = PatternFill("solid", fgColor="1E2A32")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")

    for c in containers:
        sm = stage_meta(c.current_stage_code) or {}
        row = [
            c.ref_no, c.container_no or "", c.container_type or "",
            c.client.name if c.client else "",
            _route(c),
            sm.get("title_ru", c.current_stage_code),
            days_on_stage(c), "Да" if is_stuck(c) else "",
        ]
        if show_money:
            m = _money(c)
            row += [
                round(m["income"]), round(m["expense"]), round(m["margin"]),
                round(m["paid"]), round(m["outstanding"]),
                "Да" if m["outstanding"] > 0 else "",
            ]
        ws.append(row)

    # авто-ширина + денежные колонки числом
    for col in ws.columns:
        width = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(width + 3, 42)

    # строка ИТОГО по деньгам
    if show_money and containers:
        totals = [_money(c) for c in containers]
        total_row = [""] * len(_BASE_HEAD)
        total_row[0] = "ИТОГО"
        total_row += [
            round(sum(t["income"] for t in totals)),
            round(sum(t["expense"] for t in totals)),
            round(sum(t["margin"] for t in totals)),
            round(sum(t["paid"] for t in totals)),
            round(sum(t["outstanding"] for t in totals)),
            "",
        ]
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=containers.xlsx"},
    )
